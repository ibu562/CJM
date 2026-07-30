# -*- coding: utf-8 -*-
"""Одноразовый импорт: CJM_матрица_решений_встреча.xlsx -> persona-map/data/*.csv
Запуск из корня проекта:  python persona-map/tools/import_from_xlsx.py
После импорта канон данных живёт в CSV; сайт собирается build_data.py."""
import openpyxl, csv, os, re, sys

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(os.path.dirname(ROOT), "Deliverables", "CJM_матрица_решений_встреча.xlsx")
DATA = os.path.join(ROOT, "data")
os.makedirs(DATA, exist_ok=True)

wb = openpyxl.load_workbook(XLSX, data_only=True)

PERSONA_IDS = ["P1", "P2", "P3", "P4", "P5"]
PERSONA_NAMES = {"P1": "Дмитрий", "P2": "Марина", "P3": "Елена", "P4": "Пётр", "P5": "Александра"}
# Оси числами для scatter-графика (1=низкая … 3.5=максимальная; зрелость 1=новичок … 3=продвинутая)
AXES_NUM = {
    "P1": (2.5, 2.0, 1.0),
    "P2": (3.0, 2.0, 1.0),
    "P3": (2.0, 3.5, 1.5),
    "P4": (2.5, 1.5, 2.0),
    "P5": (3.5, 1.0, 3.0),
}
DOORS = {"P1": "School €100/мес", "P2": "School / маршрут €600", "P3": "Полис €50", "P4": "курс €220", "P5": "School €100 · LTV max"}
SIZES = {"P1": 30, "P2": 30, "P3": 15, "P4": 15, "P5": 10}

def cell(ws, r, c):
    v = ws.cell(r, c).value
    return str(v).strip() if v is not None else ""

# ---------- 1. personas.csv ----------
ws = wb["Персоны (профили)"]
FIELD_MAP = {
    "Тип": "type", "Market size": "size", "Background": "background",
    "Demographics": "demographics", "Личность и интересы": "personality",
    "Goals": "goals", "Motivations": "motivations", "Expectations": "expectations",
    "Frustrations": "frustrations", "Skills Tips": "skills", "Quote": "quote",
    "Сегмент по запросу": "olga", "Оси": "axes_text", "Экономика": "economy",
    "Достоверность": "confidence", "Сценарий карты": "scenario",
}
profiles = {pid: {} for pid in PERSONA_IDS}
for r in range(2, ws.max_row + 1):
    section = cell(ws, r, 1)
    if not section:
        continue
    key = next((v for k, v in FIELD_MAP.items() if section.startswith(k)), None)
    if not key:
        print("!! пропущена секция профиля:", section); continue
    for i, pid in enumerate(PERSONA_IDS):
        profiles[pid][key] = cell(ws, r, 2 + i)

with open(os.path.join(DATA, "personas.csv"), "w", encoding="utf-8-sig", newline="") as f:
    w = csv.writer(f)
    cols = ["id", "name", "door", "size_pct", "axis_teacher", "axis_community", "axis_maturity",
            "type", "size", "background", "demographics", "personality", "goals", "motivations",
            "expectations", "frustrations", "skills", "quote", "olga", "axes_text", "economy",
            "confidence", "scenario"]
    w.writerow(cols)
    for pid in PERSONA_IDS:
        p = profiles[pid]
        t, c_, m = AXES_NUM[pid]
        w.writerow([pid, PERSONA_NAMES[pid], DOORS[pid], SIZES[pid], t, c_, m] +
                   [p.get(k, "") for k in cols[7:]])
print("personas.csv:", len(PERSONA_IDS), "персон,", len(profiles["P1"]), "полей")

# ---------- 2. stages.csv (заголовки листа CJM P1 + Цель бизнеса) ----------
ws = wb["CJM P1"]
MACRO = {**{i: "AWARE" for i in (1, 2, 3)}, **{i: "JOIN" for i in (4, 5, 6, 7, 8)},
         **{i: "USE" for i in (9, 10)}, **{i: "DEVELOP" for i in (11, 12)}, 13: "ADVOCATE"}
stages = []
# найти строки заголовков и "Цель бизнеса" по содержимому
hdr_row = next(r for r in range(1, ws.max_row + 1) if cell(ws, r, 1).startswith("Секция"))
bg_row = next(r for r in range(1, ws.max_row + 1) if cell(ws, r, 1) == "Цель бизнеса")
for c in range(2, 15):
    h = cell(ws, hdr_row, c)  # "1 Триггер [ОРБИТА]"
    m = re.match(r"(\d+)\s+(.*?)\s*\[(.*?)\]", h)
    sid, title, zone = int(m.group(1)), m.group(2), m.group(3)
    stages.append({"id": sid, "title": title, "zone": zone, "macro": MACRO[sid],
                   "phase": "САЙТ" if sid <= 8 else "ПЛАТФОРМА",
                   "business_goal": cell(ws, bg_row, c)})
with open(os.path.join(DATA, "stages.csv"), "w", encoding="utf-8-sig", newline="") as f:
    w = csv.DictWriter(f, fieldnames=["id", "title", "zone", "macro", "phase", "business_goal"])
    w.writeheader(); w.writerows(stages)
print("stages.csv:", len(stages), "этапов")

# ---------- 3. cjm.csv (5 карт × 13 этапов) ----------
SECTION_MAP = {
    "Цель персоны": "goal", "Действие": "action", "Каналы и точки контакта": "channels",
    "Боли": "pains", "Эмоция (Плутчик)": "emotion", "A-HA · Момент истины": "aha",
    "Данные": "data", "Голос": "voice", "Возможности (TO-BE)": "opportunities",
    "Задачи (TO-BE)": "tasks",
}
def parse_emotion(text):
    """ '+1 · Interest (интерес)' -> (1, 'Interest (интерес)'); '−2 · Rage … — ДНО' -> (-2, …) """
    t = text.replace("−", "-").replace("–", "-")
    m = re.match(r"^\s*([+-]?\d)\s*[·:]?\s*(.*)$", t, re.S)
    if m:
        return int(m.group(1)), m.group(2).strip()
    return "", text

# Вилки эмоций (значения веток из контент-пака, 🟢): (persona, stage) -> (основная, вторая ветка)
FORKS = {("P1", 12): (1, -1), ("P2", 12): (1, -1), ("P3", 9): (2, -2)}

rows = []
for pid in PERSONA_IDS:
    ws = wb[f"CJM {pid}"]
    smap = {}
    for r in range(1, ws.max_row + 1):
        sec = cell(ws, r, 1)
        key = SECTION_MAP.get(sec)
        if key:
            smap[key] = r
    missing = set(SECTION_MAP.values()) - set(smap)
    if missing:
        print(f"!! {pid}: нет секций {missing}"); sys.exit(1)
    for c in range(2, 15):
        sid = c - 1
        rec = {"persona_id": pid, "stage_id": sid}
        for key, r in smap.items():
            rec[key] = cell(ws, r, c)
        ev, et = parse_emotion(rec.pop("emotion"))
        ev2 = ""
        if (pid, sid) in FORKS:
            ev, ev2 = FORKS[(pid, sid)]
        rec["emotion_value"], rec["emotion_value2"], rec["emotion_term"] = ev, ev2, et
        rows.append(rec)
with open(os.path.join(DATA, "cjm.csv"), "w", encoding="utf-8-sig", newline="") as f:
    cols = ["persona_id", "stage_id", "goal", "action", "channels", "pains",
            "emotion_value", "emotion_value2", "emotion_term", "aha", "data", "voice", "opportunities", "tasks"]
    w = csv.DictWriter(f, fieldnames=cols)
    w.writeheader(); w.writerows(rows)
bad = [r for r in rows if r["emotion_value"] == ""]
print("cjm.csv:", len(rows), "ячеек карт; эмоций не распарсено:", len(bad),
      [(b["persona_id"], b["stage_id"]) for b in bad])

# ---------- 4. matrix.csv (статусы из заливки) ----------
ws = wb["Матрица персона × этап"]
COLOR_STATUS = {"FCE5A6": "yellow", "C6E0B4": "green", "F5B7B1": "red",
                "F8CBAD": "orange", "EDEDED": "gray"}
mrows = []
for r in range(3, 16):  # 13 этапов
    sid = int(re.match(r"(\d+)", cell(ws, r, 1)).group(1))
    for i, pid in enumerate(PERSONA_IDS):
        c = ws.cell(r, 2 + i)
        rgb = (c.fill.start_color.rgb or "") if c.fill and c.fill.start_color else ""
        status = COLOR_STATUS.get(str(rgb)[-6:], "")
        mrows.append({"persona_id": pid, "stage_id": sid, "status": status,
                      "text": str(c.value).strip() if c.value else ""})
    mrows.append({"persona_id": "ALL", "stage_id": sid, "status": "",
                  "text": cell(ws, r, 7)})
with open(os.path.join(DATA, "matrix.csv"), "w", encoding="utf-8-sig", newline="") as f:
    w = csv.DictWriter(f, fieldnames=["persona_id", "stage_id", "status", "text"])
    w.writeheader(); w.writerows(mrows)
nostatus = [(m["persona_id"], m["stage_id"]) for m in mrows if m["persona_id"] != "ALL" and not m["status"]]
print("matrix.csv:", len(mrows), "строк; без статуса:", nostatus)
print("\nГотово. Канон данных: persona-map/data/*.csv")
